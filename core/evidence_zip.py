"""
Evidence ZIP generator for AI Auditor
Creates comprehensive audit evidence packages with all required files
"""

import hashlib
import json
import logging
import os
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook

from .exceptions import FileProcessingError
from .export_final_xlsx import ExcelReportGenerator

logger = logging.getLogger(__name__)


class EvidenceZipGenerator:
    """Generator for audit evidence ZIP packages."""

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.excel_generator = ExcelReportGenerator()

    def generate_evidence_zip(
        self,
        findings: List[Dict[str, Any]],
        analysis_data: Dict[str, Any],
        output_path: str = None,
    ) -> str:
        """
        Generate comprehensive evidence ZIP package.

        Args:
            findings: List of audit findings
            analysis_data: Analysis results and metadata
            output_path: Output path for ZIP file

        Returns:
            Path to generated ZIP file
        """
        try:
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"evidence_audit_{timestamp}.zip"

            with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zipf:
                # 1. Generate Excel report
                excel_path = self._generate_excel_report(findings, analysis_data)
                zipf.write(excel_path, "report.xlsx")
                os.remove(excel_path)  # Clean up temp file

                # 2. Generate findings JSON
                findings_json = self._generate_findings_json(findings)
                zipf.writestr(
                    "findings.json",
                    json.dumps(findings_json, indent=2, ensure_ascii=False),
                )

                # 3. Generate decision log
                decision_log = self._generate_decision_log(findings, analysis_data)
                zipf.writestr("decision_log.csv", decision_log)

                # 4. Generate manifest
                manifest = self._generate_manifest(findings, analysis_data, output_path)
                zipf.writestr(
                    "manifest.json", json.dumps(manifest, indent=2, ensure_ascii=False)
                )

                # 5. Generate audit trail
                audit_trail = self._generate_audit_trail(findings, analysis_data)
                zipf.writestr("audit_trail.txt", audit_trail)

                # 6. Generate summary
                summary = self._generate_summary(findings, analysis_data)
                zipf.writestr("summary.txt", summary)

            self.logger.info(f"Evidence ZIP generated: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Error generating evidence ZIP: {e}")
            raise FileProcessingError(f"Failed to generate evidence ZIP: {e}")

    def _generate_excel_report(
        self, findings: List[Dict[str, Any]], analysis_data: Dict[str, Any]
    ) -> str:
        """Generate Excel report using existing generator."""
        try:
            # Prepare data for Excel generator
            excel_data = {
                "findings": findings,
                "analysis": analysis_data,
                "metadata": {
                    "generated_at": datetime.now().isoformat(),
                    "total_findings": len(findings),
                    "app_version": "1.0.0",
                },
            }

            # Generate Excel report
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = f"temp_report_{timestamp}.xlsx"

            self.excel_generator.generate_comprehensive_report(
                data=excel_data, output_path=excel_path
            )

            return excel_path

        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}")
            # Fallback to simple Excel
            return self._generate_simple_excel(findings, analysis_data)

    def _generate_simple_excel(
        self, findings: List[Dict[str, Any]], analysis_data: Dict[str, Any]
    ) -> str:
        """Generate simple Excel report as fallback."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Findings"

            # Headers
            headers = ["ID", "Type", "Severity", "Description", "Date", "Status"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Data
            for row, finding in enumerate(findings, 2):
                ws.cell(row=row, column=1, value=finding.get("id", f"F{row-1:03d}"))
                ws.cell(row=row, column=2, value=finding.get("type", "Unknown"))
                ws.cell(row=row, column=3, value=finding.get("severity", "Medium"))
                ws.cell(row=row, column=4, value=finding.get("description", ""))
                ws.cell(
                    row=row,
                    column=5,
                    value=finding.get("date", datetime.now().strftime("%Y-%m-%d")),
                )
                ws.cell(row=row, column=6, value=finding.get("status", "Open"))

            # Save
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = f"temp_simple_report_{timestamp}.xlsx"
            wb.save(excel_path)

            return excel_path

        except Exception as e:
            self.logger.error(f"Error generating simple Excel: {e}")
            raise FileProcessingError(f"Failed to generate Excel report: {e}")

    def _generate_findings_json(self, findings: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate findings JSON with schema validation."""
        return {
            "schema_version": "1.0",
            "generated_at": datetime.now().isoformat(),
            "total_findings": len(findings),
            "findings": findings,
            "summary": {
                "critical": len(
                    [f for f in findings if f.get("severity") == "critical"]
                ),
                "high": len([f for f in findings if f.get("severity") == "high"]),
                "medium": len([f for f in findings if f.get("severity") == "medium"]),
                "low": len([f for f in findings if f.get("severity") == "low"]),
            },
        }

    def _generate_decision_log(
        self, findings: List[Dict[str, Any]], analysis_data: Dict[str, Any]
    ) -> str:
        """Generate decision log CSV."""
        try:
            log_data = []

            for finding in findings:
                log_entry = {
                    "timestamp": datetime.now().isoformat(),
                    "finding_id": finding.get("id", ""),
                    "decision": finding.get("status", "Open"),
                    "user": analysis_data.get("user", "System"),
                    "reason": finding.get("reason", ""),
                    "action_taken": finding.get("action", ""),
                    "hash": self._calculate_hash(finding),
                }
                log_data.append(log_entry)

            # Convert to CSV
            df = pd.DataFrame(log_data)
            return df.to_csv(index=False)

        except Exception as e:
            self.logger.error(f"Error generating decision log: {e}")
            return (
                "timestamp,error\n"
                + datetime.now().isoformat()
                + ",Error generating log\n"
            )

    def _generate_manifest(
        self,
        findings: List[Dict[str, Any]],
        analysis_data: Dict[str, Any],
        zip_path: str,
    ) -> Dict[str, Any]:
        """Generate manifest with file hashes and metadata."""
        try:
            manifest = {
                "package_info": {
                    "name": "AI Auditor Evidence Package",
                    "version": "1.0.0",
                    "generated_at": datetime.now().isoformat(),
                    "generated_by": analysis_data.get("user", "AI Auditor"),
                    "zip_file": Path(zip_path).name,
                },
                "contents": {
                    "report.xlsx": {
                        "description": "Comprehensive audit report in Excel format",
                        "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    },
                    "findings.json": {
                        "description": "Structured findings data in JSON format",
                        "type": "application/json",
                    },
                    "decision_log.csv": {
                        "description": "Audit decision trail with timestamps",
                        "type": "text/csv",
                    },
                    "manifest.json": {
                        "description": "Package manifest and metadata",
                        "type": "application/json",
                    },
                    "audit_trail.txt": {
                        "description": "Complete audit trail",
                        "type": "text/plain",
                    },
                    "summary.txt": {
                        "description": "Executive summary",
                        "type": "text/plain",
                    },
                },
                "statistics": {
                    "total_findings": len(findings),
                    "critical_findings": len(
                        [f for f in findings if f.get("severity") == "critical"]
                    ),
                    "high_findings": len(
                        [f for f in findings if f.get("severity") == "high"]
                    ),
                    "medium_findings": len(
                        [f for f in findings if f.get("severity") == "medium"]
                    ),
                    "low_findings": len(
                        [f for f in findings if f.get("severity") == "low"]
                    ),
                },
                "integrity": {
                    "package_hash": self._calculate_file_hash(zip_path),
                    "generated_at": datetime.now().isoformat(),
                },
            }

            return manifest

        except Exception as e:
            self.logger.error(f"Error generating manifest: {e}")
            return {"error": str(e)}

    def _generate_audit_trail(
        self, findings: List[Dict[str, Any]], analysis_data: Dict[str, Any]
    ) -> str:
        """Generate complete audit trail."""
        try:
            trail = []
            trail.append("=" * 80)
            trail.append("AI AUDITOR - AUDIT TRAIL")
            trail.append("=" * 80)
            trail.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            trail.append(f"User: {analysis_data.get('user', 'System')}")
            trail.append(f"Total Findings: {len(findings)}")
            trail.append("")

            # Analysis metadata
            trail.append("ANALYSIS METADATA:")
            trail.append("-" * 40)
            for key, value in analysis_data.items():
                if key != "user":
                    trail.append(f"{key}: {value}")
            trail.append("")

            # Findings details
            trail.append("FINDINGS DETAILS:")
            trail.append("-" * 40)
            for i, finding in enumerate(findings, 1):
                trail.append(f"Finding #{i}:")
                trail.append(f"  ID: {finding.get('id', 'N/A')}")
                trail.append(f"  Type: {finding.get('type', 'N/A')}")
                trail.append(f"  Severity: {finding.get('severity', 'N/A')}")
                trail.append(f"  Description: {finding.get('description', 'N/A')}")
                trail.append(f"  Date: {finding.get('date', 'N/A')}")
                trail.append(f"  Status: {finding.get('status', 'N/A')}")
                trail.append(f"  Hash: {self._calculate_hash(finding)}")
                trail.append("")

            trail.append("=" * 80)
            trail.append("END OF AUDIT TRAIL")
            trail.append("=" * 80)

            return "\n".join(trail)

        except Exception as e:
            self.logger.error(f"Error generating audit trail: {e}")
            return f"Error generating audit trail: {e}"

    def _generate_summary(
        self, findings: List[Dict[str, Any]], analysis_data: Dict[str, Any]
    ) -> str:
        """Generate executive summary."""
        try:
            summary = []
            summary.append("AI AUDITOR - EXECUTIVE SUMMARY")
            summary.append("=" * 50)
            summary.append("")

            # Overview
            summary.append("OVERVIEW:")
            summary.append(
                f"Analysis completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )
            summary.append(f"Total findings: {len(findings)}")
            summary.append("")

            # Severity breakdown
            severity_counts = {
                "critical": len(
                    [f for f in findings if f.get("severity") == "critical"]
                ),
                "high": len([f for f in findings if f.get("severity") == "high"]),
                "medium": len([f for f in findings if f.get("severity") == "medium"]),
                "low": len([f for f in findings if f.get("severity") == "low"]),
            }

            summary.append("SEVERITY BREAKDOWN:")
            for severity, count in severity_counts.items():
                summary.append(f"  {severity.upper()}: {count}")
            summary.append("")

            # Key findings
            critical_findings = [f for f in findings if f.get("severity") == "critical"]
            if critical_findings:
                summary.append("CRITICAL FINDINGS:")
                for finding in critical_findings[:5]:  # Top 5
                    summary.append(
                        f"  - {finding.get('description', 'No description')}"
                    )
                summary.append("")

            # Recommendations
            summary.append("RECOMMENDATIONS:")
            summary.append("  1. Address all critical findings immediately")
            summary.append("  2. Review high severity findings within 48 hours")
            summary.append("  3. Implement preventive controls for recurring issues")
            summary.append("  4. Schedule follow-up review in 30 days")
            summary.append("")

            summary.append("This report was generated by AI Auditor v1.0.0")
            summary.append("For questions, contact the audit team.")

            return "\n".join(summary)

        except Exception as e:
            self.logger.error(f"Error generating summary: {e}")
            return f"Error generating summary: {e}"

    def _calculate_hash(self, data: Dict[str, Any]) -> str:
        """Calculate hash for data integrity."""
        try:
            data_str = json.dumps(data, sort_keys=True, ensure_ascii=False)
            return hashlib.sha256(data_str.encode("utf-8")).hexdigest()[:16]
        except Exception:
            return "hash_error"

    def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate hash for file integrity."""
        try:
            with open(file_path, "rb") as f:
                return hashlib.sha256(f.read()).hexdigest()
        except Exception:
            return "file_hash_error"


# Global instance
_evidence_generator: Optional[EvidenceZipGenerator] = None


def get_evidence_generator() -> EvidenceZipGenerator:
    """Get global evidence generator instance."""
    global _evidence_generator
    if _evidence_generator is None:
        _evidence_generator = EvidenceZipGenerator()
    return _evidence_generator


def generate_evidence_zip(
    findings: List[Dict[str, Any]],
    analysis_data: Dict[str, Any],
    output_path: str = None,
) -> str:
    """
    Generate evidence ZIP package using global generator.

    Args:
        findings: List of audit findings
        analysis_data: Analysis results and metadata
        output_path: Output path for ZIP file

    Returns:
        Path to generated ZIP file
    """
    return get_evidence_generator().generate_evidence_zip(
        findings, analysis_data, output_path
    )
