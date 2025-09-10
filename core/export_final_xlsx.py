"""
Production Excel report generation system.
Creates comprehensive audit reports with formulas and formatting.
"""

import logging
from pathlib import Path
from typing import Any, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .exceptions import FileProcessingError


class ExcelReportGenerator:
    """Production Excel report generator with advanced formatting."""

    def __init__(self):
        self.logger = logging.getLogger(__name__)

        # Color scheme for professional reports
        self.colors = {
            "header": "366092",
            "subheader": "4F81BD",
            "success": "92D050",
            "warning": "FFC000",
            "error": "FF0000",
            "neutral": "D9D9D9",
            "light_blue": "B7D7E8",
            "light_green": "C6EFCE",
            "light_red": "FFC7CE",
        }

        # Fonts
        self.fonts = {
            "header": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
            "subheader": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
            "normal": Font(name="Calibri", size=11),
            "bold": Font(name="Calibri", size=11, bold=True),
            "small": Font(name="Calibri", size=10),
        }

        # Borders
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def _apply_header_style(self, ws, row: int, cols: int, title: str):
        """Apply header styling to a range."""
        # Merge cells for title
        ws.merge_cells(f"A{row}:{chr(65 + cols - 1)}{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = self.fonts["header"]
        ws[f"A{row}"].fill = PatternFill(
            start_color=self.colors["header"],
            end_color=self.colors["header"],
            fill_type="solid",
        )
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

        return row + 1

    def _apply_table_style(
        self, ws, start_row: int, end_row: int, start_col: int, end_col: int
    ):
        """Apply table styling to a range."""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border

                # Alternate row colors
                if row % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=self.colors["light_blue"],
                        end_color=self.colors["light_blue"],
                        fill_type="solid",
                    )

    def _create_summary_sheet(self, wb: Workbook, summary_data: Dict[str, Any]) -> None:
        """Create executive summary sheet."""
        ws = wb.create_sheet("Executive Summary", 0)

        row = 1
        row = self._apply_header_style(ws, row, 4, "AI Auditor - Executive Summary")
        row += 1

        # Summary statistics
        ws[f"A{row}"] = "Total Invoices Processed:"
        ws[f"B{row}"] = summary_data.get("total_invoices", 0)
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 1

        ws[f"A{row}"] = "Successfully Matched:"
        ws[f"B{row}"] = summary_data.get("matched_invoices", 0)
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 1

        ws[f"A{row}"] = "Unmatched Invoices:"
        ws[f"B{row}"] = summary_data.get("unmatched_invoices", 0)
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 1

        ws[f"A{row}"] = "Match Rate:"
        ws[f"B{row}"] = f"={ws[f'B{row-2}'].coordinate}/{ws[f'B{row-3}'].coordinate}"
        ws[f"B{row}"].number_format = "0.00%"
        ws[f"A{row}"].font = self.fonts["bold"]
        row += 2

        # Risk indicators
        row = self._apply_header_style(ws, row, 4, "Risk Indicators")
        row += 1

        risk_indicators = summary_data.get("risk_indicators", {})
        for risk, value in risk_indicators.items():
            ws[f"A{row}"] = risk
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = self.fonts["bold"]
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

    def _create_detailed_results_sheet(
        self, wb: Workbook, results_df: pd.DataFrame
    ) -> None:
        """Create detailed results sheet."""
        ws = wb.create_sheet("Detailed Results")

        row = 1
        row = self._apply_header_style(
            ws, row, len(results_df.columns), "Detailed Invoice Matching Results"
        )
        row += 1

        # Add headers
        for col, header in enumerate(results_df.columns, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.fonts["subheader"]
            cell.fill = PatternFill(
                start_color=self.colors["subheader"],
                end_color=self.colors["subheader"],
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center")

        row += 1

        # Add data
        for _, data_row in results_df.iterrows():
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = self.fonts["normal"]

                # Color coding based on status
                if col == 1 and str(value).lower() == "znaleziono":
                    cell.fill = PatternFill(
                        start_color=self.colors["light_green"],
                        end_color=self.colors["light_green"],
                        fill_type="solid",
                    )
                elif col == 1 and str(value).lower() == "brak":
                    cell.fill = PatternFill(
                        start_color=self.colors["light_red"],
                        end_color=self.colors["light_red"],
                        fill_type="solid",
                    )

            row += 1

        # Apply table styling
        self._apply_table_style(ws, 2, row - 1, 1, len(results_df.columns))

        # Adjust column widths
        for col in range(1, len(results_df.columns) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _create_top50_mismatches_sheet(
        self, wb: Workbook, mismatches_df: pd.DataFrame
    ) -> None:
        """Create Top 50 mismatches sheet."""
        ws = wb.create_sheet("Top 50 Mismatches")

        row = 1
        row = self._apply_header_style(
            ws, row, len(mismatches_df.columns), "Top 50 Invoice Mismatches"
        )
        row += 1

        # Add headers
        for col, header in enumerate(mismatches_df.columns, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.fonts["subheader"]
            cell.fill = PatternFill(
                start_color=self.colors["subheader"],
                end_color=self.colors["subheader"],
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center")

        row += 1

        # Add data (limit to top 50)
        top_50 = mismatches_df.head(50)
        for _, data_row in top_50.iterrows():
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = self.fonts["normal"]
                cell.fill = PatternFill(
                    start_color=self.colors["light_red"],
                    end_color=self.colors["light_red"],
                    fill_type="solid",
                )

            row += 1

        # Apply table styling
        self._apply_table_style(ws, 2, row - 1, 1, len(mismatches_df.columns))

        # Adjust column widths
        for col in range(1, len(mismatches_df.columns) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20

    def _create_statistics_sheet(
        self, wb: Workbook, stats_data: Dict[str, Any]
    ) -> None:
        """Create statistics and analysis sheet."""
        ws = wb.create_sheet("Statistics & Analysis")

        row = 1
        row = self._apply_header_style(ws, row, 4, "Statistical Analysis")
        row += 1

        # Match criteria breakdown
        row = self._apply_header_style(ws, row, 2, "Match Criteria Breakdown")
        row += 1

        criteria_stats = stats_data.get("match_criteria", {})
        for criteria, count in criteria_stats.items():
            ws[f"A{row}"] = criteria
            ws[f"B{row}"] = count
            ws[f"A{row}"].font = self.fonts["bold"]
            row += 1

        row += 1

        # Confidence distribution
        row = self._apply_header_style(ws, row, 2, "Confidence Distribution")
        row += 1

        confidence_stats = stats_data.get("confidence_distribution", {})
        for range_str, count in confidence_stats.items():
            ws[f"A{row}"] = range_str
            ws[f"B{row}"] = count
            ws[f"A{row}"].font = self.fonts["bold"]
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

    def generate_report(
        self,
        all_invoices_df: pd.DataFrame,
        verdicts_df: pd.DataFrame,
        summary_data: Dict[str, Any],
        output_path: Path,
    ) -> None:
        """Generate comprehensive Excel report."""
        try:
            self.logger.info(f"Generating Excel report: {output_path}")

            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create summary sheet
            self._create_summary_sheet(wb, summary_data)

            # Create detailed results sheet
            self._create_detailed_results_sheet(wb, verdicts_df)

            # Create top 50 mismatches sheet
            mismatches = verdicts_df[verdicts_df["status"] == "brak"].copy()
            if not mismatches.empty:
                self._create_top50_mismatches_sheet(wb, mismatches)

            # Create statistics sheet
            stats_data = self._calculate_statistics(verdicts_df)
            self._create_statistics_sheet(wb, stats_data)

            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Excel report saved successfully: {output_path}")

        except Exception as e:
            raise FileProcessingError(f"Failed to generate Excel report: {e}")

    def _calculate_statistics(self, verdicts_df: pd.DataFrame) -> Dict[str, Any]:
        """Calculate statistics from verdicts data."""
        stats = {}

        # Match criteria breakdown
        if "criteria" in verdicts_df.columns:
            criteria_counts = verdicts_df["criteria"].value_counts().to_dict()
            stats["match_criteria"] = criteria_counts

        # Confidence distribution
        if "confidence" in verdicts_df.columns:
            confidence_ranges = {
                "0.0 - 0.2": 0,
                "0.2 - 0.4": 0,
                "0.4 - 0.6": 0,
                "0.6 - 0.8": 0,
                "0.8 - 1.0": 0,
            }

            for conf in verdicts_df["confidence"]:
                if conf < 0.2:
                    confidence_ranges["0.0 - 0.2"] += 1
                elif conf < 0.4:
                    confidence_ranges["0.2 - 0.4"] += 1
                elif conf < 0.6:
                    confidence_ranges["0.4 - 0.6"] += 1
                elif conf < 0.8:
                    confidence_ranges["0.6 - 0.8"] += 1
                else:
                    confidence_ranges["0.8 - 1.0"] += 1

            stats["confidence_distribution"] = confidence_ranges

        return stats


# Global instance
_excel_generator = ExcelReportGenerator()


def generate_excel_report(
    all_invoices_df: pd.DataFrame,
    verdicts_df: pd.DataFrame,
    summary_data: Dict[str, Any],
    output_path: Path,
) -> None:
    """Convenience function to generate Excel report."""
    _excel_generator.generate_report(
        all_invoices_df, verdicts_df, summary_data, output_path
    )
